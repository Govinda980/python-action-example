import openpyxl


def read_multiple_sheet():
    path = 'dummy_data.xlsx'
    wb = openpyxl.load_workbook(path)
    sheet_data_dict = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_data = []
        for row in ws.iter_rows(values_only=True):
            sheet_data.append(row)

        sheet_data_dict[ws] = sheet_data
    print(sheet_data_dict)


def write_into_multiple_sheet():
    sheet_data_dic = {
        'Employees': [
            ('Name', 'Age', 'Salary', 'Role'),
            ('Alice', 30, 70000, 'Engineer'),
            ('Bob', 25, 50000, 'Designer'),
            ('Charlie', 35, 80000, 'Manager')
        ],
        'Doctors': [
            ('Name', 'Experience', 'Salary', 'Specialization'),
            ('Riyansh', 1, 100000, 'Pediatrics'),
            ('Omendra', 2, 150000, 'General'),
            ('Anish', 3, 200000, 'Cardiology')
        ],
        'Cricketers': [
            ('Name', 'Age', 'Matches', 'Role'),
            ('Anik', 13, 50, 'Batsman'),
            ('Manish', 4, 20, 'Bowler'),
            ('Nitesh', 13, 30, 'All-rounder')
        ]
    }
    wb = openpyxl.Workbook()
    for sheet, data in sheet_data_dic.items():
        ws = wb.create_sheet(title=sheet)
        for row_index, row_data in enumerate(data, start=1):
            for col_index, col_data in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index, value=col_data)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save('output_multiple_sheet.xlsx')


if __name__ == '__main__':
    # read_multiple_sheet()
    write_into_multiple_sheet()
