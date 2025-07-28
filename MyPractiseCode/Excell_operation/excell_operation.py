import openpyxl


def read_excel_openpyxl():
    try:
        # Load the workbook and select the active sheet
        employees = []
        path = 'dummy_data.xlsx'  # Replace with your actual file path
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            employees.append(row)
        print(employees)
    except Exception as e:
        print(e)


def write_to_excel_openpyxl():
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Defining some dummy data to write
    data = [('Name', 'Age', 'Salary', 'Role'), ('Alice', 30, 70000, 'Engineer'), ('Bob', 25, 50000, 'Designer'),
            ('Charlie', 35, 80000, 'Manager'), ('David', 28, 60000, 'Developer'), ('Riyansh', 1, 100000, 'Doctor'),
            ('Omendra', 22, 100000, 'Doctor'), ('Anik', 13, 200000, 'Cricketer'), ('Anish', 10, 60000, 'Suba'),
            ('Manish', 4, 234550, 'ASI'), ('Nitesh', 13, 2000000, 'HA')]

    # Writing data to the worksheet
    for row_index, row_data in enumerate(data, start=1):
        for col_index, col_value in enumerate(row_data, start=1):
            ws.cell(row=row_index, column=col_index, value=col_value)

    # Save the workbook to a file
    ws.save('output_data.xlsx')


if __name__ == '__main__':
    # write_to_excel_openpyxl()
    read_excel_openpyxl()
