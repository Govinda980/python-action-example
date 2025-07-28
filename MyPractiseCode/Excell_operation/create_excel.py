import openpyxl
def create_excell_with_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active

    sheet1 = wb.create_sheet(title="A")
    sheet2 = wb.create_sheet(title="B")
    print(wb.sheetnames)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save("create_excell_test.xlsx")


if __name__ == '__main__':
    create_excell_with_sheet()